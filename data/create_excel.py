"""
生成 hospital_bed_usage_data_2.xlsx 基础数据
Sheet1: 医院科室信息（地区、医院id、医院名称、科室id、科室名称、病房id、病房名称、病房总床数）
Sheet2: 床位信息（病房id、床位号）
"""

import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

HOSPITALS = [
    {"id": "H001", "name": "玛丽医院", "region": "香港岛", "floors": 9},
    {"id": "H002", "name": "东区尤德夫人那打素医院", "region": "香港岛", "floors": 10},
    {"id": "H003", "name": "伊利沙伯医院", "region": "九龙", "floors": 8},
    {"id": "H004", "name": "广华医院", "region": "九龙", "floors": 7},
    {"id": "H005", "name": "威尔斯亲王医院", "region": "新界", "floors": 9},
    {"id": "H006", "name": "屯门医院", "region": "新界", "floors": 8},
]

DEPARTMENTS = [
    "内科", "外科", "骨科", "儿科", "妇产科",
    "心脏科", "神经科", "肿瘤科",
]

WARD_NAMES = {
    "内科": ["内科A病房", "内科B病房"],
    "外科": ["外科A病房", "外科B病房"],
    "骨科": ["骨科A病房", "骨科B病房"],
    "儿科": ["儿科病房"],
    "妇产科": ["妇产科A病房", "妇产科B病房"],
    "心脏科": ["心脏科A病房", "心脏科B病房"],
    "神经科": ["神经科A病房", "神经科B病房"],
    "肿瘤科": ["肿瘤科A病房", "肿瘤科B病房", "肿瘤科C病房"],
}

# 每个病房的床位数范围
BEDS_PER_WARD = (15, 30)

OUTPUT_DIR = os.path.dirname(__file__)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "hospital_bed_usage_data_2.xlsx")


def generate_excel():
    wb = Workbook()

    # ── Sheet 1: 医院科室信息 ──
    ws1 = wb.active
    ws1.title = "医院科室信息"
    headers1 = ["地区", "医院id", "医院名称", "科室id", "科室名称", "病房id", "病房名称", "病房总床数"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    dept_counter = 0
    ward_counter = 0
    bed_data = []  # (ward_id, bed_number)

    for hosp in HOSPITALS:
        for dept_name in DEPARTMENTS:
            dept_counter += 1
            dept_id = f"D{dept_counter:03d}"
            ward_names = WARD_NAMES.get(dept_name, ["普通病房"])
            for wn in ward_names:
                ward_counter += 1
                ward_id = f"W{ward_counter:04d}"
                n_beds = random.randint(*BEDS_PER_WARD)

                ws1.cell(row=row_idx, column=1, value=hosp["region"])
                ws1.cell(row=row_idx, column=2, value=hosp["id"])
                ws1.cell(row=row_idx, column=3, value=hosp["name"])
                ws1.cell(row=row_idx, column=4, value=dept_id)
                ws1.cell(row=row_idx, column=5, value=dept_name)
                ws1.cell(row=row_idx, column=6, value=ward_id)
                ws1.cell(row=row_idx, column=7, value=wn)
                ws1.cell(row=row_idx, column=8, value=n_beds)

                for b in range(1, n_beds + 1):
                    bed_data.append((ward_id, b))

                row_idx += 1

    # ── Sheet 2: 床位信息 ──
    ws2 = wb.create_sheet("床位信息")
    headers2 = ["病房id", "床位号"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, (ward_id, bed_no) in enumerate(bed_data, 2):
        ws2.cell(row=i, column=1, value=ward_id)
        ws2.cell(row=i, column=2, value=bed_no)

    # 列宽
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # 保存
    wb.save(OUTPUT_FILE)
    total_wards = ward_counter
    total_beds = len(bed_data)
    print(f"✅ Excel 已生成: {OUTPUT_FILE}")
    print(f"   医院: {len(HOSPITALS)} 所")
    print(f"   科室: {dept_counter} 个 (每院 {len(DEPARTMENTS)} 个)")
    print(f"   病房: {total_wards} 间")
    print(f"   床位: {total_beds} 张")
    return OUTPUT_FILE


if __name__ == "__main__":
    seed = 42
    random.seed(seed)
    generate_excel()
